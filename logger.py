import openpyxl
from datetime import datetime

# Define header constant
HEADER = ['User', 'Activity', 'Date', 'Time']

def log_activity(user, activity):
    current_datetime = datetime.now()
    current_hour = current_datetime.hour
    current_minute = current_datetime.minute
    current_second = current_datetime.second

    today_date = current_datetime.strftime("%m/%d/%y")
    today_time = f"{current_hour:02d}:{current_minute:02d}:{current_second:02d}"

    log_entry = {
        'User': user,
        'Activity': activity,
        'Date': today_date,
        'Time': today_time
    }

    file_path = 'databases/activity_log.xlsx'

    # Write the log entry to Excel
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    worksheet = workbook.active

    # Write header if the file is empty
    if worksheet.max_row == 1:
        worksheet.append(HEADER)

    # Write the log entry
    log_data = [log_entry[col] for col in HEADER]
    worksheet.append(log_data)

    # Adjust column widths
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        adjusted_width = (length + 2) * 1.2  # Add some padding and adjust width
        worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    workbook.save(file_path)
